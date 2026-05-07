"""
job_hunter_agent/client_tracker.py
=====================================
Saves every business lead and email status to:
  - data/clients.json  (fast lookup / dedup)
  - data/clients.xlsx  (Excel spreadsheet — open with LibreOffice/Excel)

Columns in Excel:
  Name | Category | City | Website | Email | Phone |
  Subject | Status | Date Found | Date Emailed | Notes
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

# Supabase sync (graceful — works even if not configured)
try:
    from supabase_sync import upsert_lead as _sb_upsert, is_configured as _sb_ok
    SUPABASE_ENABLED = True
except ImportError:
    SUPABASE_ENABLED = False
    _sb_ok = lambda: False
    _sb_upsert = lambda *a, **k: None

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)

JSON_FILE = DATA_DIR / "clients.json"

# ── Try openpyxl for Excel ──────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
    EXCEL_FILE = DATA_DIR / "clients.xlsx"
except ImportError:
    EXCEL_OK = False
    logger.warning("openpyxl not installed — Excel export disabled. Run: pip install openpyxl")


COLUMNS = [
    "name", "category", "city", "website", "email", "phone",
    "contact_name", "email_subject", "status", "date_found",
    "date_emailed", "source", "used_ai", "notes",
]

STATUS_COLORS = {
    "found":       "FFFFFF",   # white
    "no_email":    "FFF3CD",   # yellow
    "email_sent":  "D4EDDA",   # green
    "bounced":     "F8D7DA",   # red
    "replied":     "CCE5FF",   # blue
    "converted":   "E2D9F3",   # purple
}


# ─────────────────────────────────────────────────────────────
#  LOAD / SAVE JSON
# ─────────────────────────────────────────────────────────────
def _load_clients() -> Dict[str, Dict]:
    """Load clients dict keyed by normalized name."""
    if JSON_FILE.exists():
        try:
            return json.loads(JSON_FILE.read_text(encoding="utf-8"))
        except Exception as e:
            logger.error(f"Failed to load clients.json: {e}")
    return {}


def _save_clients(clients: Dict[str, Dict]):
    JSON_FILE.write_text(json.dumps(clients, indent=2, ensure_ascii=False), encoding="utf-8")


def _key(name: str) -> str:
    return name.lower().strip()


# ─────────────────────────────────────────────────────────────
#  EXCEL
# ─────────────────────────────────────────────────────────────
def _rebuild_excel(clients: Dict[str, Dict]):
    """Rebuild the Excel file from scratch from current JSON data."""
    if not EXCEL_OK:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Client Leads"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [c.replace("_", " ").title() for c in COLUMNS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 25

    # Data rows
    for row_idx, client in enumerate(clients.values(), start=2):
        row_color = STATUS_COLORS.get(client.get("status", "found"), "FFFFFF")
        fill = PatternFill("solid", fgColor=row_color)

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = client.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    col_widths = {
        "A": 30, "B": 14, "C": 16, "D": 35, "E": 32, "F": 16,
        "G": 20, "H": 40, "I": 14, "J": 14, "K": 14, "L": 16,
        "M": 10, "N": 30,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(EXCEL_FILE)
    logger.debug(f"Excel updated: {EXCEL_FILE}")


# ─────────────────────────────────────────────────────────────
#  PUBLIC API
# ─────────────────────────────────────────────────────────────
def is_already_contacted(name: str) -> bool:
    """Return True if we already sent an email to this business."""
    clients = _load_clients()
    client = clients.get(_key(name))
    if not client:
        return False
    return client.get("status") in ("email_sent", "replied", "converted")


def save_client_lead(lead: Dict) -> bool:
    """
    Save a new business lead. Returns True if newly added, False if already exists.
    """
    clients = _load_clients()
    key = _key(lead.get("name", ""))
    if not key:
        return False

    if key in clients:
        # Update any new fields
        existing = clients[key]
        for field in ["email", "phone", "website", "contact_name"]:
            if lead.get(field) and not existing.get(field):
                existing[field] = lead[field]
        clients[key] = existing
        _save_clients(clients)
        return False

    # New lead
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    clients[key] = {
        "name": lead.get("name", ""),
        "category": lead.get("category", ""),
        "city": lead.get("city", ""),
        "website": lead.get("website", ""),
        "email": lead.get("email", ""),
        "phone": lead.get("phone", ""),
        "contact_name": lead.get("contact_name", ""),
        "email_subject": "",
        "status": "found",
        "date_found": now,
        "date_emailed": "",
        "source": lead.get("source", ""),
        "used_ai": "",
        "notes": "",
    }
    _save_clients(clients)
    _rebuild_excel(clients)
    return True


def mark_email_sent(name: str, subject: str, used_ai: bool = False):
    """Update lead status after sending email. Auto-syncs to Supabase."""
    clients = _load_clients()
    key = _key(name)
    if key in clients:
        clients[key]["status"] = "email_sent"
        clients[key]["email_subject"] = subject
        clients[key]["date_emailed"] = datetime.now().isoformat()
        clients[key]["used_ai"] = str(used_ai)
        _save_clients(clients)
        _rebuild_excel(clients)
        # Auto-sync to Supabase cloud
        if SUPABASE_ENABLED and _sb_ok():
            try:
                _sb_upsert(key, clients[key])
            except Exception:
                pass


def update_client_status(name: str, status: str, notes: str = ""):
    """Update status: found / no_email / email_sent / bounced / replied / converted."""
    clients = _load_clients()
    key = _key(name)
    if key in clients:
        clients[key]["status"] = status
        if notes:
            clients[key]["notes"] = notes
        _save_clients(clients)
        _rebuild_excel(clients)


def update_client_contact(name: str, email: str = "", phone: str = "", contact_name: str = ""):
    """Add contact details discovered by contact_finder."""
    clients = _load_clients()
    key = _key(name)
    if key in clients:
        if email:
            clients[key]["email"] = email
        if phone:
            clients[key]["phone"] = phone
        if contact_name:
            clients[key]["contact_name"] = contact_name
        _save_clients(clients)


def get_stats() -> Dict:
    """Return summary counts."""
    clients = _load_clients()
    stats: Dict[str, int] = {}
    for c in clients.values():
        s = c.get("status", "found")
        stats[s] = stats.get(s, 0) + 1
    stats["total"] = len(clients)
    return stats


def get_all_clients() -> List[Dict]:
    return list(_load_clients().values())


# ─────────────────────────────────────────────────────────────
#  QUICK TEST
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    test = {
        "name": "City General Hospital",
        "category": "hospital",
        "city": "Amsterdam",
        "website": "https://cityhospital.com",
        "email": "info@cityhospital.com",
        "phone": "+31 20 123 4567",
        "contact_name": "Dr. Smith",
        "source": "google_maps",
    }
    added = save_client_lead(test)
    print(f"Added: {added}")
    mark_email_sent("City General Hospital", "Custom Hospital CRM for City General Hospital", used_ai=True)
    print(f"Stats: {get_stats()}")
    print(f"Already contacted: {is_already_contacted('City General Hospital')}")
